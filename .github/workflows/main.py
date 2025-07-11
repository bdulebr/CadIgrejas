# main.py - Sistema de Gestão de Visitantes, Membros e Funcionários
# Autor: Especialista Python
# Versão: 2.0 (Aprimorada)

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, Toplevel, simpledialog
except ModuleNotFoundError:
    raise ImportError("O módulo 'tkinter' não está disponível. Verifique se ele está instalado e compatível com seu ambiente Python.")

import openpyxl
from datetime import datetime
import os
import hashlib
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import matplotlib.pyplot as plt
from collections import Counter

# --- ARQUIVOS E CONSTANTES ---
VISITORS_FILE = "visitantes.xlsx"
MEMBERS_FILE = "membros.xlsx"
EMPLOYEES_FILE = "funcionarios.xlsx"
USERS_FILE = "usuarios.xlsx"
LOG_FILE = "logs.xlsx"

VISITORS_HEADERS = ["ID", "Nome", "Telefone", "Email", "Endereço", "Data Nascimento", "Data Visita", "Igreja", "Observações", "Família"]
MEMBERS_HEADERS = ["ID", "Nome", "Telefone", "Email", "Data Nascimento", "Cargo", "Batizado", "Endereço", "Família"]
EMPLOYEES_HEADERS = ["ID", "Nome", "Cargo", "Telefone", "Email", "Data Admissão", "Salário", "Observações", "Família"]
USERS_HEADERS = ["Usuário", "SenhaHash", "Permissão"]
LOG_HEADERS = ["Usuário", "Ação", "DataHora"]
PERMISSOES = ["admin", "completo", "leitura"]

# --- UTILITÁRIOS E SETUP ---

def hash_password(pw):
    """Gera um hash SHA256 para a senha fornecida."""
    return hashlib.sha256(pw.encode()).hexdigest()

def criar_arquivos_se_nao_existir():
    """Cria os arquivos .xlsx necessários com cabeçalhos se eles não existirem."""
    arquivos = [
        (VISITORS_FILE, VISITORS_HEADERS),
        (MEMBERS_FILE, MEMBERS_HEADERS),
        (EMPLOYEES_FILE, EMPLOYEES_HEADERS),
        (LOG_FILE, LOG_HEADERS)
    ]
    for file, headers in arquivos:
        if not os.path.exists(file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(headers)
            wb.save(file)
            
    # Cria arquivo de usuários com admin padrão se não existir
    if not os.path.exists(USERS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(USERS_HEADERS)
        ws.append(["admin", hash_password("admin"), "admin"])
        wb.save(USERS_FILE)

def log_action(usuario, acao):
    """Registra uma ação do usuário no arquivo de log."""
    try:
        wb = openpyxl.load_workbook(LOG_FILE)
        ws = wb.active
        ws.append([usuario, acao, datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
        wb.save(LOG_FILE)
    except FileNotFoundError:
        messagebox.showwarning("Log Erro", "Arquivo de log não encontrado. A ação não será registrada.")

# --- CLASSES DA INTERFACE GRÁFICA ---

class LoginWindow:
    """Janela de login inicial."""
    def __init__(self, root, on_success):
        self.root = root
        self.on_success = on_success
        self.win = tk.Frame(root, bg='#2e2e2e')
        self.win.pack(expand=True, fill='both')

        style = ttk.Style()
        style.configure("TLabel", background="#2e2e2e", foreground="white")
        style.configure("TButton", padding=6, relief="flat", background="#4a4a4a", foreground="white")
        style.map("TButton", background=[('active', '#6a6a6a')])

        frame = ttk.Frame(self.win, padding=40, style="TFrame")
        frame.place(relx=0.5, rely=0.5, anchor="center")

        ttk.Label(frame, text="Sistema de Gestão da Igreja", font=("Segoe UI", 16, "bold")).pack(pady=(0, 20))
        
        ttk.Label(frame, text="Usuário:", font=("Segoe UI", 12)).pack(pady=5, anchor='w')
        self.user_entry = ttk.Entry(frame, width=35, font=("Segoe UI", 11))
        self.user_entry.pack()

        ttk.Label(frame, text="Senha:", font=("Segoe UI", 12)).pack(pady=5, anchor='w')
        self.pass_entry = ttk.Entry(frame, width=35, show="*", font=("Segoe UI", 11))
        self.pass_entry.pack()
        self.pass_entry.bind("<Return>", self.attempt_login)

        ttk.Button(frame, text="Entrar", command=self.attempt_login, width=15).pack(pady=20)

    def attempt_login(self, event=None):
        usuario = self.user_entry.get()
        senha = self.pass_entry.get()
        
        try:
            wb = openpyxl.load_workbook(USERS_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == usuario and row[1] == hash_password(senha):
                    log_action(usuario, "Login bem-sucedido")
                    permissao = row[2]
                    self.win.destroy()
                    self.on_success(usuario, permissao)
                    return
            messagebox.showerror("Erro de Login", "Usuário ou senha inválidos.", parent=self.root)
        except FileNotFoundError:
            messagebox.showerror("Erro Crítico", f"Arquivo de usuários '{USERS_FILE}' não encontrado.", parent=self.root)

class MainMenu:
    """Menu principal da aplicação."""
    def __init__(self, root, usuario, permissao):
        self.root = root
        self.usuario = usuario
        self.permissao = permissao

        self.frame = ttk.Frame(self.root, padding=20)
        self.frame.pack(expand=True, fill="both")

        ttk.Label(self.frame, text=f"Bem-vindo, {usuario}!", font=("Segoe UI", 16, "bold")).pack(pady=10)
        ttk.Label(self.frame, text=f"Permissão: {permissao.capitalize()}", font=("Segoe UI", 10)).pack(pady=(0, 20))

        btn_width = 35
        ttk.Button(self.frame, text="Cadastro de Visitantes", width=btn_width, command=self.abrir_visitantes).pack(pady=5)
        ttk.Button(self.frame, text="Cadastro de Membros", width=btn_width, command=self.abrir_membros).pack(pady=5)
        ttk.Button(self.frame, text="Cadastro de Funcionários", width=btn_width, command=self.abrir_funcionarios).pack(pady=5)
        ttk.Button(self.frame, text="Relatórios em PDF", width=btn_width, command=self.gerar_pdf).pack(pady=5)
        ttk.Button(self.frame, text="Relatórios Gráficos", width=btn_width, command=self.gerar_grafico).pack(pady=5)
        
        if self.permissao == "admin":
             ttk.Button(self.frame, text="Gerenciar Usuários", width=btn_width, command=self.gerenciar_usuarios).pack(pady=5)
        
        ttk.Button(self.frame, text="Sair (Logout)", width=btn_width, command=self.logout, style="Accent.TButton").pack(pady=20)

    def abrir_janela_dados(self, titulo, arquivo, cabecalhos):
        DataWindow(self.root, titulo, arquivo, cabecalhos, self.permissao)

    def abrir_visitantes(self):
        self.abrir_janela_dados("Cadastro de Visitantes", VISITORS_FILE, VISITORS_HEADERS)

    def abrir_membros(self):
        self.abrir_janela_dados("Cadastro de Membros", MEMBERS_FILE, MEMBERS_HEADERS)

    def abrir_funcionarios(self):
        self.abrir_janela_dados("Cadastro de Funcionários", EMPLOYEES_FILE, EMPLOYEES_HEADERS)
    
    def gerenciar_usuarios(self):
        self.abrir_janela_dados("Gerenciamento de Usuários", USERS_FILE, USERS_HEADERS)

    def gerar_pdf(self):
        try:
            wb = openpyxl.load_workbook(VISITORS_FILE)
            ws = wb.active
            pdf = canvas.Canvas("relatorio_visitantes.pdf", pagesize=A4)
            pdf.setTitle("Relatório de Visitantes")
            pdf.drawString(72, 800, "Relatório de Visitantes")
            y = 750
            for row in ws.iter_rows(min_row=2, values_only=True):
                texto = ", ".join(str(c) if c is not None else "" for c in row)
                pdf.drawString(72, y, texto)
                y -= 20
                if y < 40:
                    pdf.showPage()
                    y = 800
            pdf.save()
            messagebox.showinfo("Sucesso", "Relatório PDF 'relatorio_visitantes.pdf' gerado.")
        except Exception as e:
            messagebox.showerror("Erro PDF", f"Não foi possível gerar o PDF: {e}")

    def gerar_grafico(self):
        try:
            wb = openpyxl.load_workbook(VISITORS_FILE)
            ws = wb.active
            # Coluna 8 = Igreja (índice 7)
            igrejas = [row[7] for row in ws.iter_rows(min_row=2, values_only=True) if row[7]]
            if not igrejas:
                messagebox.showinfo("Gráfico", "Não há dados de igrejas de origem para gerar o gráfico.")
                return
            contagem = Counter(igrejas)
            plt.figure(figsize=(10, 6))
            plt.bar(contagem.keys(), contagem.values(), color='#4a90e2')
            plt.title("Visitantes por Igreja de Origem")
            plt.ylabel("Número de Visitantes")
            plt.xticks(rotation=45, ha="right")
            plt.tight_layout()
            plt.savefig("grafico_igrejas.png")
            messagebox.showinfo("Sucesso", "Gráfico 'grafico_igrejas.png' salvo com sucesso.")
            plt.show()
        except Exception as e:
            messagebox.showerror("Erro Gráfico", f"Não foi possível gerar o gráfico: {e}")

    def logout(self):
        log_action(self.usuario, "Logout")
        for widget in self.root.winfo_children():
            widget.destroy()
        LoginWindow(self.root, lambda u, p: MainMenu(self.root, u, p))

class DataWindow:
    """Janela genérica para visualização e manipulação de dados (CRUD)."""
    def __init__(self, root, title, filename, headers, permissao):
        if permissao == 'leitura':
            self.read_only = True
        else:
            self.read_only = False
            if title == "Gerenciamento de Usuários" and permissao != 'admin':
                messagebox.showerror("Acesso Negado", "Você não tem permissão para gerenciar usuários.")
                return

        self.win = Toplevel(root)
        self.win.title(title)
        self.win.geometry("1000x600")
        self.filename = filename
        self.headers = headers

        # --- Frames ---
        top_frame = ttk.Frame(self.win, padding=10)
        top_frame.pack(fill='x')
        
        tree_frame = ttk.Frame(self.win, padding=10)
        tree_frame.pack(expand=True, fill='both')

        # --- Widgets de Pesquisa e Botões ---
        ttk.Label(top_frame, text="Pesquisar:").pack(side='left', padx=(0, 5))
        self.search_entry = ttk.Entry(top_frame, width=40)
        self.search_entry.pack(side='left', fill='x', expand=True)
        self.search_entry.bind("<KeyRelease>", self.filter_tree)

        if not self.read_only:
            ttk.Button(top_frame, text="Adicionar", command=self.add_record).pack(side='left', padx=5)
            ttk.Button(top_frame, text="Editar", command=self.edit_record).pack(side='left', padx=5)
            ttk.Button(top_frame, text="Excluir", command=self.delete_record).pack(side='left', padx=5)

        # --- TreeView para exibir dados ---
        self.tree = ttk.Treeview(tree_frame, columns=self.headers, show='headings')
        for col in self.headers:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')
        self.tree.pack(expand=True, fill='both')

        self.load_data()

    def load_data(self):
        """Carrega os dados do arquivo Excel e popula o TreeView."""
        for i in self.tree.get_children():
            self.tree.delete(i)
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Ocultar hash de senha na UI
                if self.filename == USERS_FILE:
                    display_row = list(row)
                    display_row[1] = '********'
                    self.tree.insert("", "end", values=display_row)
                else:
                    self.tree.insert("", "end", values=row)
        except FileNotFoundError:
            messagebox.showerror("Erro", f"Arquivo '{self.filename}' não encontrado.", parent=self.win)

    def filter_tree(self, event=None):
        """Filtra os dados no TreeView com base na pesquisa."""
        search_term = self.search_entry.get().lower()
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            match = any(search_term in str(cell).lower() for cell in row)
            if match:
                if self.filename == USERS_FILE:
                    display_row = list(row)
                    display_row[1] = '********'
                    self.tree.insert("", "end", values=display_row)
                else:
                    self.tree.insert("", "end", values=row)

    def add_record(self):
        """Abre um diálogo para adicionar um novo registro."""
        dialog = RecordDialog(self.win, "Adicionar Registro", self.headers, filename=self.filename)
        if dialog.result:
            try:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                
                # Gerar novo ID
                new_id = 1
                if ws.max_row > 1:
                    last_id = ws.cell(row=ws.max_row, column=1).value
                    if last_id is not None and str(last_id).isdigit():
                        new_id = int(last_id) + 1

                new_row = [new_id] + dialog.result
                
                # Hash da senha se for cadastro de usuário
                if self.filename == USERS_FILE:
                    new_row[2] = hash_password(new_row[2]) # Senha é o segundo campo do resultado
                
                ws.append(new_row)
                wb.save(self.filename)
                self.load_data()
            except Exception as e:
                messagebox.showerror("Erro ao Salvar", f"Não foi possível adicionar o registro: {e}", parent=self.win)

    def edit_record(self):
        """Abre um diálogo para editar o registro selecionado."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Nenhuma Seleção", "Por favor, selecione um registro para editar.", parent=self.win)
            return
        
        # Obter valores reais do arquivo, não da TreeView (que mascara a senha)
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        record_id = self.tree.item(selected_item[0])['values'][0]
        
        original_values = None
        row_to_edit = -1
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == record_id:
                original_values = row
                row_to_edit = i
                break
        
        if not original_values:
            messagebox.showerror("Erro", "Registro não encontrado no arquivo.", parent=self.win)
            return

        dialog = RecordDialog(self.win, "Editar Registro", self.headers, initial_data=original_values, filename=self.filename)
        if dialog.result:
            try:
                updated_row = [record_id] + dialog.result
                # Se for usuário e a senha foi alterada, hasheia a nova
                if self.filename == USERS_FILE and dialog.result[1] != '********':
                     updated_row[2] = hash_password(dialog.result[1])

                for col_num, value in enumerate(updated_row, start=1):
                    ws.cell(row=row_to_edit, column=col_num, value=value)
                
                wb.save(self.filename)
                self.load_data()
            except Exception as e:
                messagebox.showerror("Erro ao Salvar", f"Não foi possível editar o registro: {e}", parent=self.win)

    def delete_record(self):
        """Exclui o registro selecionado."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Nenhuma Seleção", "Por favor, selecione um registro para excluir.", parent=self.win)
            return

        if messagebox.askyesno("Confirmar Exclusão", "Tem certeza que deseja excluir o registro selecionado?", parent=self.win):
            try:
                wb = openpyxl.load_workbook(self.filename)
                ws = wb.active
                record_id = self.tree.item(selected_item[0])['values'][0]

                row_to_delete = -1
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value == record_id:
                        row_to_delete = i
                        break
                
                if row_to_delete != -1:
                    ws.delete_rows(row_to_delete)
                    wb.save(self.filename)
                    self.load_data()
                else:
                    messagebox.showerror("Erro", "Não foi possível encontrar o registro para excluir.", parent=self.win)
            except Exception as e:
                messagebox.showerror("Erro ao Excluir", f"Não foi possível excluir o registro: {e}", parent=self.win)

class RecordDialog(Toplevel):
    """Diálogo genérico para adicionar ou editar um registro."""
    def __init__(self, parent, title, headers, initial_data=None, filename=None):
        super().__init__(parent)
        self.title(title)
        self.geometry("400x500")
        self.transient(parent)
        self.grab_set()

        self.headers = headers[1:] # Ignora o ID
        self.initial_data = initial_data[1:] if initial_data else None
        self.filename = filename
        self.result = None
        self.entries = {}

        frame = ttk.Frame(self, padding=20)
        frame.pack(expand=True, fill="both")

        for i, header in enumerate(self.headers):
            ttk.Label(frame, text=f"{header}:").grid(row=i, column=0, sticky='w', pady=2)
            
            # Campo de senha especial para usuários
            if self.filename == USERS_FILE and header == "SenhaHash":
                entry = ttk.Entry(frame, show="*")
                if self.initial_data:
                    entry.insert(0, "********") # Placeholder
            elif self.filename == USERS_FILE and header == "Permissão":
                entry = ttk.Combobox(frame, values=PERMISSOES, state="readonly")
            else:
                entry = ttk.Entry(frame)
            
            if self.initial_data:
                # Não exibir hash da senha
                if not (self.filename == USERS_FILE and header == "SenhaHash"):
                    entry.insert(0, self.initial_data[i] if self.initial_data[i] is not None else "")
            
            entry.grid(row=i, column=1, sticky='ew', pady=2)
            self.entries[header] = entry

        frame.grid_columnconfigure(1, weight=1)

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=len(self.headers), column=0, columnspan=2, pady=15)
        ttk.Button(btn_frame, text="Salvar", command=self.on_ok).pack(side='left', padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=self.destroy).pack(side='left', padx=10)

        self.wait_window(self)

    def on_ok(self):
        self.result = []
        for header in self.headers:
            value = self.entries[header].get()
            # Validação simples
            if not value:
                # Permite senha em branco na edição (significa não alterar)
                is_editing_user_password = (self.filename == USERS_FILE and
                                            header == "SenhaHash" and
                                            self.initial_data and
                                            value == "********")
                if not is_editing_user_password:
                    messagebox.showwarning("Campo Vazio", f"O campo '{header}' não pode estar vazio.", parent=self)
                    return
            self.result.append(value)
        
        self.destroy()

# --- FUNÇÃO PRINCIPAL ---

def main():
    """Função principal que inicia a aplicação."""
    criar_arquivos_se_nao_existir()
    
    root = tk.Tk()
    root.title("Sistema de Gestão - Palavra de Vida Enseada")
    root.geometry("600x500")
    
    # Estilo da aplicação
    style = ttk.Style(root)
    style.theme_use('clam')
    style.configure("TButton", padding=10, font=("Segoe UI", 10))
    style.configure("Accent.TButton", foreground="white", background="#c0392b")
    style.map("Accent.TButton", background=[('active', '#e74c3c')])

    def start_main_menu(usuario, permissao):
        for widget in root.winfo_children():
            widget.destroy()
        MainMenu(root, usuario, permissao)

    LoginWindow(root, start_main_menu)
    root.mainloop()

if __name__ == "__main__":
    main()
