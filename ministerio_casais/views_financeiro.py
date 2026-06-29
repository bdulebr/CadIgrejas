from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from permissoes.decorators import requer_permissao
from django.contrib import messages
from django.utils import timezone
from .models import PagamentoCursoCasal, PagamentoInscricaoEvento, DespesaMinisterio, EventoCasal
from django.db.models import Sum
from datetime import datetime, date

@login_required
@requer_permissao('casais', 'ver')
def painel_financeiro(request):
    mes_atual = timezone.now().month
    ano_atual = timezone.now().year

    # Filtros
    data_inicio_str = request.GET.get('data_inicio')
    data_fim_str = request.GET.get('data_fim')

    if data_inicio_str and data_fim_str:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
    else:
        # Default: mês atual
        data_inicio = date(ano_atual, mes_atual, 1)
        # Último dia do mês (hacky approach)
        import calendar
        _, last_day = calendar.monthrange(ano_atual, mes_atual)
        data_fim = date(ano_atual, mes_atual, last_day)

    # Entradas de Cursos
    pgtos_cursos = PagamentoCursoCasal.objects.filter(data_pagamento__date__gte=data_inicio, data_pagamento__date__lte=data_fim)
    total_cursos = pgtos_cursos.aggregate(total=Sum('valor_pago'))['total'] or 0

    # Entradas de Eventos
    pgtos_eventos = PagamentoInscricaoEvento.objects.filter(data_pagamento__date__gte=data_inicio, data_pagamento__date__lte=data_fim)
    total_eventos = pgtos_eventos.aggregate(total=Sum('valor_pago'))['total'] or 0

    total_receitas = total_cursos + total_eventos

    # Despesas
    despesas = DespesaMinisterio.objects.filter(data_despesa__gte=data_inicio, data_despesa__lte=data_fim).order_by('-data_despesa')
    total_despesas = despesas.aggregate(total=Sum('valor'))['total'] or 0

    lucro = total_receitas - total_despesas

    eventos_ativos = EventoCasal.objects.filter(status__in=['Aberto', 'Planejamento', 'Lotado'])

    context = {
        'data_inicio': data_inicio.strftime('%Y-%m-%d'),
        'data_fim': data_fim.strftime('%Y-%m-%d'),
        'total_cursos': total_cursos,
        'total_eventos': total_eventos,
        'total_receitas': total_receitas,
        'total_despesas': total_despesas,
        'lucro': lucro,
        'despesas': despesas,
        'eventos_ativos': eventos_ativos,
        'pgtos_cursos': pgtos_cursos.order_by('-data_pagamento'),
        'pgtos_eventos': pgtos_eventos.order_by('-data_pagamento')
    }

    return render(request, 'ministerio_casais/financeiro/painel.html', context)

@login_required
@requer_permissao('casais', 'editar')
def nova_despesa(request):
    if request.method == 'POST':
        descricao = request.POST.get('descricao')
        valor = request.POST.get('valor')
        categoria = request.POST.get('categoria')
        data_despesa = request.POST.get('data_despesa')
        evento_id = request.POST.get('evento_id')

        evento = EventoCasal.objects.filter(id=evento_id).first() if evento_id else None

        if valor and data_despesa:
            try:
                valor_f = float(valor.replace(',', '.'))
                despesa = DespesaMinisterio.objects.create(
                    descricao=descricao,
                    valor=valor_f,
                    categoria=categoria,
                    data_despesa=data_despesa,
                    evento_vinculado=evento
                )
                if request.FILES.get('comprovante'):
                    despesa.comprovante = request.FILES.get('comprovante')
                    despesa.save()
                messages.success(request, 'Despesa registrada com sucesso!')
            except ValueError:
                messages.error(request, 'Valor inválido.')

    return redirect('painel_financeiro_casais')

@login_required
@requer_permissao('casais', 'ver')
def gerar_relatorio_tesouraria(request):
    if request.method == 'POST':
        data_inicio_str = request.POST.get('data_inicio')
        data_fim_str = request.POST.get('data_fim')
        formato = request.POST.get('formato')
        enviar_email = request.POST.get('enviar_email')
        email_destino = request.POST.get('email_destino')

        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()

        pgtos_cursos = PagamentoCursoCasal.objects.filter(data_pagamento__date__gte=data_inicio, data_pagamento__date__lte=data_fim)
        pgtos_eventos = PagamentoInscricaoEvento.objects.filter(data_pagamento__date__gte=data_inicio, data_pagamento__date__lte=data_fim)
        despesas = DespesaMinisterio.objects.filter(data_despesa__gte=data_inicio, data_despesa__lte=data_fim)

        total_rec = (pgtos_cursos.aggregate(t=Sum('valor_pago'))['t'] or 0) + (pgtos_eventos.aggregate(t=Sum('valor_pago'))['t'] or 0)
        total_desp = despesas.aggregate(t=Sum('valor'))['t'] or 0
        lucro = total_rec - total_desp

        # Gera PDF
        if formato == 'pdf' or enviar_email == 'on':
            from io import BytesIO
            from django.http import HttpResponse
            from django.template.loader import render_to_string
            import xhtml2pdf.pisa as pisa

            html_str = render_to_string('ministerio_casais/financeiro/pdf_relatorio.html', {
                'data_inicio': data_inicio,
                'data_fim': data_fim,
                'pgtos_cursos': pgtos_cursos,
                'pgtos_eventos': pgtos_eventos,
                'despesas': despesas,
                'total_rec': total_rec,
                'total_desp': total_desp,
                'lucro': lucro,
                'data_geracao': timezone.now()
            })

            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html_str.encode("UTF-8")), result)
            if not pdf.err:
                pdf_bytes = result.getvalue()

                # Se for email
                if enviar_email == 'on' and email_destino:
                    from django.core.mail import EmailMultiAlternatives
                    from django.conf import settings
                    msg = EmailMultiAlternatives(
                        subject=f"Relatório Financeiro Casais ({data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')})",
                        body="Segue em anexo o relatório financeiro do Ministério de Casais.",
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[email_destino]
                    )
                    msg.attach(f"relatorio_casais_{data_inicio.strftime('%Y%m%d')}.pdf", pdf_bytes, 'application/pdf')
                    msg.send()
                    messages.success(request, f'Relatório enviado para {email_destino}!')
                    return redirect('painel_financeiro_casais')

                # Download
                if formato == 'pdf':
                    response = HttpResponse(pdf_bytes, content_type='application/pdf')
                    response['Content-Disposition'] = f'attachment; filename="relatorio_financeiro_{data_inicio.strftime("%Y%m%d")}.pdf"'
                    return response

        # Gera XLSX
        if formato == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font
            from django.http import HttpResponse

            wb = openpyxl.Workbook()

            # Aba Resumo
            ws = wb.active
            ws.title = "Resumo"
            ws['A1'] = "Relatório Financeiro - Ministério de Casais"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A3'] = f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"

            ws.append([])
            ws.append(["Métrica", "Valor (R$)"])
            ws.append(["Receitas - Cursos", float(pgtos_cursos.aggregate(t=Sum('valor_pago'))['t'] or 0)])
            ws.append(["Receitas - Eventos", float(pgtos_eventos.aggregate(t=Sum('valor_pago'))['t'] or 0)])
            ws.append(["Total Receitas", float(total_rec)])
            ws.append(["Total Despesas", float(total_desp)])
            ws.append(["Lucro / Prejuízo", float(lucro)])

            # Aba Entradas
            ws2 = wb.create_sheet(title="Entradas (Cursos e Eventos)")
            ws2.append(["Data", "Tipo", "Origem", "Forma Pgto", "Valor"])
            for p in pgtos_cursos:
                ws2.append([p.data_pagamento.strftime('%d/%m/%Y'), "Curso", f"{p.matricula.casal.nomes_juntos} ({p.matricula.turma.curso.nome})", p.forma_pagamento, float(p.valor_pago)])
            for p in pgtos_eventos:
                ws2.append([p.data_pagamento.strftime('%d/%m/%Y'), "Evento", f"{p.inscricao.casal.nomes_juntos} ({p.inscricao.evento.titulo})", p.forma_pagamento, float(p.valor_pago)])

            # Aba Saidas
            ws3 = wb.create_sheet(title="Despesas")
            ws3.append(["Data", "Categoria", "Descrição", "Evento Vinculado", "Valor"])
            for d in despesas:
                ws3.append([d.data_despesa.strftime('%d/%m/%Y'), d.categoria, d.descricao, d.evento_vinculado.titulo if d.evento_vinculado else "-", float(d.valor)])

            from io import BytesIO
            response_io = BytesIO()
            wb.save(response_io)

            if enviar_email == 'on' and email_destino:
                from django.core.mail import EmailMultiAlternatives
                from django.conf import settings
                msg = EmailMultiAlternatives(
                    subject=f"Relatório Financeiro Casais ({data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')})",
                    body="Segue em anexo a planilha do relatório financeiro do Ministério de Casais.",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[email_destino]
                )
                msg.attach(f"relatorio_casais_{data_inicio.strftime('%Y%m%d')}.xlsx", response_io.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                msg.send()
                messages.success(request, f'Planilha enviada para {email_destino}!')
                return redirect('painel_financeiro_casais')

            response = HttpResponse(response_io.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="relatorio_financeiro_{data_inicio.strftime("%Y%m%d")}.xlsx"'
            return response

    return redirect('painel_financeiro_casais')
