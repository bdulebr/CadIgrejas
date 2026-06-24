"""
* PROJETO: Palavra de Vida Enseada - Intranet
* ARQUIVO: intranet/services/gemini_ai.py
* DESCRIÇÃO: Código-fonte do módulo
* DEV: Marcos Roberto Lira (marcos@pvenseada.org)
* VERSÃO: 0.0.1
* DATA DA ÚLTIMA ALTERAÇÃO: 16/06/2026 14:37
* LOG DE ALTERAÇÕES:
* - 16/06/2026 14:37: Auditoria e padronização global (Goal)
"""
import os
import json
import tempfile
import datetime
from google import genai
from django.conf import settings

def get_gemini_client():
    api_key = getattr(settings, 'GEMINI_API_KEY', '')
    if not api_key:
        raise Exception("Chave GEMINI_API_KEY não configurada no settings.py")
    return genai.Client(api_key=api_key)

def consultar_gemini_json(prompt_text, retries=3):
    """
    Envia um prompt para o Gemini e retorna a string de resposta, esperando um JSON puro.
    Usado pelo AI Auto-Engineer e AI Auto-Fix.
    Em caso de QuotaExceeded (429), tenta novamente com exponential backoff.
    """
    import time
    client = get_gemini_client()

    for attempt in range(retries):
        try:
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt_text
            )
            return response.text
        except Exception as e:
            if '429' in str(e) or 'RESOURCE_EXHAUSTED' in str(e):
                if attempt < retries - 1:
                    sleep_time = 20 * (attempt + 1)
                    print(f"[Gemini API] Quota Excedida (429). Aguardando {sleep_time} segundos antes da próxima tentativa...")
                    time.sleep(sleep_time)
                else:
                    raise e
            else:
                raise e

def analisar_comprovante_tesouraria(file_obj, categorias=None):
    """
    Usa o Gemini 1.5 Flash para ler uma imagem ou PDF de um comprovante/nota fiscal
    e extrair os dados em formato JSON.
    """
    # Cria arquivo temporário
    extensao = ".pdf"
    if hasattr(file_obj, 'name'):
        extensao = os.path.splitext(file_obj.name)[1].lower()

    with tempfile.NamedTemporaryFile(delete=False, suffix=extensao) as tmp:
        for chunk in file_obj.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    uploaded_file = None
    client = None
    try:
        client = get_gemini_client()

        # Faz upload via File API
        uploaded_file = client.files.upload(file=tmp_path)

        categorias_str = ""
        if categorias:
            categorias_str = "\nCategorias disponíveis no banco de dados (ID e Nome):\n"
            for c in categorias:
                categorias_str += f"ID: {c['id']} - Nome: {c['nome']}\n"

        prompt = f"""
        Você é um assistente contábil e de OCR.
        Analise o comprovante, nota fiscal ou recibo anexado.
        {categorias_str}

        Extraia as seguintes informações e retorne ESTRITAMENTE um objeto JSON válido (sem crases Markdown, sem texto antes ou depois):
        {{
          "tipo": "entrada ou saida",
          "valor": 150.50,
          "data_vencimento": "YYYY-MM-DD",
          "data_lancamento": "YYYY-MM-DDTHH:MM",
          "descricao": "Breve descrição do gasto (ex: Compra de materiais Kalunga)",
          "categoria_id": 123,
          "categoria_sugerida": "Nome da Categoria",
          "forma_pagamento": "pix, boleto, credito, debito, dinheiro, transferencia ou outros",
          "impostos": 5.50,
          "is_parcelado": true,
          "numero_parcelas": 12,
          "parcela_atual": 2
        }}

        Regras:
        1. 'tipo' deve ser 'entrada' se for um dízimo, oferta ou crédito. Deve ser 'saida' se for nota fiscal de compra, despesa, boleto pago, etc.
        2. 'valor' deve ser um número float (use ponto para separar decimais, não vírgula).
        3. 'data_vencimento' deve ser a data principal do documento no formato YYYY-MM-DD. Se não achar uma data de vencimento (ex: cupom fiscal), use a data de emissão. Se não achar data nenhuma, retorne nulo.
        4. O JSON retornado deve ser perfeitamente parseável por json.loads() em Python.
        5. Para o campo 'categoria_id', analise a despesa/receita e tente encontrar a categoria no texto "Categorias disponíveis" que mais se aproxima (Fuzzy match). Se encontrar uma correspondência razoável, preencha o campo "categoria_id" com o número inteiro do ID.
        6. Se não achar nenhuma categoria que faça sentido ou se a lista de categorias estiver vazia, defina "categoria_id" como nulo (null) e escreva um nome aproximado/inventado do que seria essa categoria no campo "categoria_sugerida" (ex: "Alimentação", "Manutenção Predial").
        7. 'forma_pagamento' deve ser categorizado dentre as opções válidas.
        8. 'impostos': se a nota indicar valor total de tributos/impostos, extraia como float. Senão, nulo.
        9. Se o documento indicar que a compra/venda é parcelada (ex: parcela 1/5), extraia 'is_parcelado' como true, 'numero_parcelas' como o total de parcelas (ex: 5) e 'parcela_atual' como a parcela daquele boleto/nota (ex: 1). Se não houver, false e nulos.
        """

        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[uploaded_file, prompt]
        )

        # Clean response
        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.endswith("```"):
            text = text[:-3]

        return json.loads(text.strip())

    except Exception as e:
        print("Erro no Gemini:", e)
        raise e
    finally:
        # Tenta excluir o arquivo do servidor do Gemini e o temporário
        if client and uploaded_file:
            try:
                client.files.delete(name=uploaded_file.name)
            except:
                pass
        try:
            os.unlink(tmp_path)
        except:
            pass

def analisar_escala_gemini(file_obj, departamentos_list, membros_list):
    """
    Função de OCR nativo (Multimodal) via Gemini com RAG embutido.
    departamentos_list: list of dicts [{'id': ..., 'nome': ...}]
    membros_list: list of dicts [{'id': ..., 'nome': ...}]
    """
    # Cria arquivo temporário
    extensao = ".pdf"
    if hasattr(file_obj, 'name'):
        extensao = os.path.splitext(file_obj.name)[1].lower()

    tmp_path = None
    with tempfile.NamedTemporaryFile(delete=False, suffix=extensao) as tmp:
        for chunk in file_obj.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    client = None
    uploaded_file = None
    try:
        client = get_gemini_client()
        uploaded_file = client.files.upload(file=tmp_path)

        dept_str = "\n".join([f"ID: {d['id']} | Nome: {d['nome']}" for d in departamentos_list])
        membros_str = "\n".join([f"ID: {m['id']} | Nome: {m['nome']}" for m in membros_list])

        prompt = f"""
        Você é um Motor Avançado de OCR para Escalas de Igreja.
        Estou anexando um arquivo de escala (geralmente um PDF ou planilha). Leia-o visual e textualmente.

        SUA MISSÃO RAG (MUITO IMPORTANTE):
        Eu vou te dar a base de dados REAL do meu sistema.
        Você só pode retornar IDs que existam nesta base de dados.
        Se um nome na escala for "Kauêzinho" e na base de dados tiver "ID: 15 | Nome: Kauê Lira (Kauêzinho)", você deve mapear para o ID 15.

        BASE DE DADOS DE DEPARTAMENTOS:
        {dept_str}

        BASE DE DADOS DE MEMBROS (ID | Nome | Apelido):
        {membros_str}

        O QUE VOCÊ DEVE RETORNAR:
        Exatamente um JSON válido, e apenas ele (sem formatadores Markdown), neste modelo:
        {{
            "departamento_id": 12, // ID do departamento que melhor bate com o título do documento
            "mes": "06", // Mês em 2 dígitos (ex: Junho -> 06)
            "ano": "2026", // Ano em 4 dígitos
            "escalas": [
                {{
                    "dia": "DD/MM/YYYY", // Data exata daquele culto
                    "turno": "manha ou noite", // Deixe vazio se não souber
                    "funcao": "Câmera", // A função escrita
                    "membros_ids": [15, 20], // Lista INTEIROS com os IDs mapeados da base de dados!
                    "observacao": "Alguma nota, caso exista"
                }}
            ]
        }}

        REGRAS VITAIS:
        1. SEPARE FUNÇÕES! Se no dia 03/06 tiver "CÂMERA: ARTHUR / GERAL: PEDRO", você deve gerar dois blocos na lista 'escalas': um para Câmera (membros_ids do Arthur) e um para Geral (membros_ids do Pedro).
        2. NÃO CRIE IDs INVENTADOS. Use somente os listados na BASE DE DADOS.
        3. BUSCA APROXIMADA (FUZZY MATCHING): O PDF pode conter apenas primeiros nomes, apelidos ou nomes com erros de digitação (ex: "Kauezinho", "Fernandinha"). Você DEVE usar inteligência e dedução aproximada para encontrar quem é a pessoa na lista de Membros fornecida, checando o primeiro nome, sobrenome e o (apelido). Mapeie para o ID correto.
        """

        import time
        retries = 3
        for attempt in range(retries):
            try:
                response = client.models.generate_content(
                    model='gemini-2.5-flash',
                    contents=[uploaded_file, prompt]
                )
                break
            except Exception as e:
                if '503' in str(e) or '429' in str(e) or 'UNAVAILABLE' in str(e):
                    if attempt < retries - 1:
                        time.sleep(3 * (attempt + 1))
                        continue
                raise e

        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:]
        if text.endswith("```"):
            text = text[:-3]
        return json.loads(text.strip())

    except Exception as e:
        print("Erro no Gemini Escalas OCR:", e)
        return []
    finally:
        if client and uploaded_file:
            try:
                client.files.delete(name=uploaded_file.name)
            except:
                pass
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except:
                pass

import openpyxl

def analisar_planilha_importacao(file_obj, categorias=None):
    """
    Lê a planilha XLSX, converte para texto, envia ao Gemini para validar
    e detectar erros críticos. Retorna JSON com os lançamentos estruturados e os erros (se houver).
    """
    client = get_gemini_client()

    # 1. Carregar XLSX
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # 2. Converter para CSV/Texto em Memória
    csv_str = ""
    for row in ws.iter_rows(values_only=True):
        row_str = []
        for cell in row:
            if cell is None:
                row_str.append("")
            else:
                # Tratar strings e remover newlines
                val = str(cell).replace('\n', ' ').replace('\r', '').replace(';', ',')
                row_str.append(val)
        csv_str += ";".join(row_str) + "\n"

    categorias_str = ""
    if categorias:
        categorias_str = "\nCategorias disponíveis (ID e Nome):\n"
        for c in categorias:
            categorias_str += f"ID: {c['id']} - Nome: {c['nome']}\n"

    prompt = f"""
    Você é um validador de importação do módulo de Tesouraria.
    Vou te passar os dados brutos extraídos de uma planilha CSV (separador ponto e vírgula).

    {categorias_str}

    DADOS DA PLANILHA:
    {csv_str}

    Sua missão é:
    1. Ignorar o cabeçalho.
    2. Identificar cada linha de lançamento e formatar os dados.
    3. Fazer o fuzzy match do nome da categoria descrita na planilha para o "categoria_id" (se não achar, nulo).
    3. Fazer o fuzzy match da categoria de acordo com a lista.
    4. Identificar data de vencimento e DATA DE LANÇAMENTO (data/hora em que a transação ocorreu de fato, no formato YYYY-MM-DDTHH:MM).
    5. Identificar impostos e parcelamentos.
    6. Identificar a forma de pagamento (pix, boleto, credito, debito, dinheiro, transferencia, outros).
    7. Validar as regras de negócio Zero-Trust: Valores não podem ser negativos, datas de vencimento devem ser coerentes (YYYY-MM-DD), o tipo DEVE ser "entrada" ou "saida".

    Se houver ERROS GRAVES (ex: valores negativos, datas absurdas como 29/02 em ano não bissexto), você deve preencher o campo "erros_críticos" no JSON com a lista de problemas encontrados para que a importação seja abortada. Se estiver tudo OK ou houver apenas warnings leves, "erros_críticos" deve ser uma lista vazia [].

    Retorne ESTRITAMENTE um objeto JSON com o formato:
    {{
      "erros_críticos": ["Erro 1", "Erro 2..."],
      "lancamentos": [
        {{
          "tipo": "entrada ou saida",
          "valor": 100.50,
          "data_vencimento": "YYYY-MM-DD",
          "data_lancamento": "YYYY-MM-DDTHH:MM",
          "descricao": "Descricao",
          "categoria_id": 123,
          "forma_pagamento": "pix",
          "impostos": 0.0,
          "is_parcelado": false,
          "numero_parcelas": 1,
          "parcela_atual": 1,
          "observacoes": ""
        }}
      ]
    }}
    Não inclua marcações Markdown.
    """

    response = client.models.generate_content(
        model='gemini-2.5-flash',
        contents=prompt
    )

    try:
        texto = response.text.strip()
        if texto.startswith("```json"):
            texto = texto[7:]
        if texto.endswith("```"):
            texto = texto[:-3]
        dados = json.loads(texto)
        return dados
    except Exception as e:
        raise Exception(f"A IA falhou em formatar o lote: {str(e)} -> Resposta original: {response.text}")

def gerar_planilha_sede_mensal(mes, ano):
    """
    1. Lê todos os lançamentos do mês e ano especificados.
    2. Lê os cabeçalhos da `planilha_padrao_sede`.
    3. Envia os cabeçalhos e os lançamentos para o Gemini mapear (adivinhar a distribuição).
    4. Recebe o JSON com as linhas mapeadas.
    5. Escreve no Excel clonado e retorna o caminho do arquivo gerado.
    """
    from tesouraria.models import Lancamento, ConfiguracaoTesouraria

    config = ConfiguracaoTesouraria.objects.first()
    if not config or not config.planilha_padrao_sede:
        raise Exception("A planilha padrão da Sede não foi configurada. Acesse Configurações para anexá-la.")

    lancamentos = Lancamento.objects.filter(data_vencimento__month=mes, data_vencimento__year=ano)
    if not lancamentos.exists():
        raise Exception(f"Nenhum lançamento encontrado para o mês {mes}/{ano}.")

    # Prepara JSON simplificado dos lançamentos
    lista_lancamentos = []
    for l in lancamentos:
        lista_lancamentos.append({
            "id": l.id,
            "tipo": l.tipo,
            "valor": float(l.valor),
            "data_vencimento": l.data_vencimento.strftime("%Y-%m-%d"),
            "data_lancamento": l.data_lancamento.strftime("%Y-%m-%dT%H:%M") if l.data_lancamento else "",
            "descricao": l.descricao,
            "categoria": l.categoria.nome if l.categoria else ""
        })

    # Lê os cabeçalhos do Excel
    template_path = config.planilha_padrao_sede.path
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Vamos assumir que o cabeçalho está na linha 1
    cabecalhos = []
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        cabecalhos.append(str(cell_val) if cell_val else f"Coluna_{col}")

    prompt = f"""
    Você é um contador experiente que precisa preencher um relatório de fechamento de caixa para a Sede da igreja.

    Abaixo estão as colunas exatas da planilha do Excel da Sede (na ordem, da esquerda para direita):
    {json.dumps(cabecalhos, ensure_ascii=False)}

    Abaixo estão os lançamentos financeiros deste mês:
    {json.dumps(lista_lancamentos, ensure_ascii=False)}

    Sua tarefa é classificar e mapear cada lançamento para as colunas corretas da planilha.

    Regras:
    1. Retorne ESTRITAMENTE um objeto JSON contendo uma lista de linhas chamada "linhas".
    2. Cada linha (objeto) deve ser um array de valores exatos correspondendo, em ordem, aos cabeçalhos acima.
       Ex: ["01/06/2026", "Entrada", "Oferta Alçada", 150.00, ...]
    3. Se uma coluna do cabeçalho for "Data", coloque a data convertida. Se for "Histórico" ou "Descrição", coloque a descrição.
    4. Se você não souber o que colocar em uma coluna, coloque vazio ("").
    5. Retorne APENAS o JSON, sem formatação markdown.

    Formato esperado:
    {{
      "linhas": [
        ["valor_col1", "valor_col2", "valor_col3"],
        ["valor_col1", "valor_col2", "valor_col3"]
      ]
    }}
    """

    client = get_gemini_client()
    response = client.models.generate_content(
        model='gemini-2.5-flash',
        contents=[prompt]
    )

    text = response.text.strip()
    if text.startswith("```json"):
        text = text[7:]
    if text.endswith("```"):
        text = text[:-3]

    try:
        dados_mapeados = json.loads(text.strip())
    except Exception as e:
        print("Erro de parse JSON:", text)
        raise Exception("A inteligência falhou em gerar os dados formatados.")

    # Começa a escrever da linha 2
    row_idx = 2
    for linha in dados_mapeados.get("linhas", []):
        for col_idx, val in enumerate(linha, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1

    # Salva o novo arquivo
    output_filename = f"Relatorio_Sede_{mes:02d}_{ano}.xlsx"
    output_path = os.path.join(tempfile.gettempdir(), output_filename)
    wb.save(output_path)

    return output_path

def gerar_escala_inteligente_gemini(departamento_nome, mes, ano, membros, eventos, regras):
    """
    Usa o Gemini 2.5 Flash para gerar uma escala inteligente, balanceada e sem conflitos.
    """
    client = get_gemini_client()

    prompt = f"""
    Você é um Motor de Inteligência Artificial de Alocação de Escalas (Workforce Management).
    Sua missão é gerar a escala do mês de {mes}/{ano} para o departamento '{departamento_nome}'.

    REGRAS DO MOTOR:
    1. Respeite estritamente os REQUISITOS (habilidades) de cada Função.
    2. Não aloque membros que estejam indisponíveis nas datas informadas.
    3. BALANCEAMENTO: Distribua a carga. Não aloque a mesma pessoa muitas vezes se houver outros disponíveis.
    4. Limite máximo geral: {regras.get('limite_mensal', 4)} vezes no mês por pessoa.
    5. Ninguém pode estar em dois lugares no mesmo dia/turno.
    6. O membro DEVE ter o ID da função ('funcao_id') listado em seu array 'funcoes_ids'. Se o membro não tiver a função associada, NÃO O ALOQUE para aquela vaga.

    MEMBROS ELEGÍVEIS (IDs, Nomes, Habilidades, Indisponibilidades):
    {json.dumps(membros, ensure_ascii=False)}

    EVENTOS DO MÊS E SLOTS NECESSÁRIOS:
    {json.dumps(eventos, ensure_ascii=False)}

    Você DEVE retornar ESTRITAMENTE o resultado no seguinte formato JSON, resolvendo o quebra-cabeça da alocação:
    {{
      "alocacoes": [
        {{
          "evento_id": "id do evento na lista",
          "data": "YYYY-MM-DD",
          "horario_inicio": "HH:MM",
          "funcao_id": "id da funcao",
          "membro_id": "ID INTEIRO do membro alocado"
        }}
      ]
    }}
    Retorne apenas o JSON, sem marcações markdown.
    """

    response = client.models.generate_content(
        model='gemini-2.5-flash',
        contents=prompt
    )

    texto = response.text.strip()
    if texto.startswith("```json"):
        texto = texto[7:]
    if texto.endswith("```"):
        texto = texto[:-3]
    dados = json.loads(texto.strip())
    return dados

def extrair_dados_membro_texto(texto_livre):
    """
    Usa o Gemini 2.5 Flash para extrair dados estruturados de um membro a partir de um texto livre.
    Retorna JSON com as chaves compatíveis com o form_perfil_mestre.html.
    """
    client = get_gemini_client()
    prompt = f"""
    Você é um assistente de Inteligência Artificial para Gestão de RH Eclesiástico.
    Sua missão é extrair as informações de um membro a partir do texto abaixo e preencher o JSON.

    Regras:
    1. Se não encontrar a informação, deixe como "".
    2. O 'sexo' deve ser "Masculino", "Feminino" ou "Outro". Tente deduzir pelo nome ou adjetivos (ex: "casado" -> Masculino).
    3. O 'estado_civil' deve ser "Solteiro(a)", "Casado(a)", "Divorciado(a)" ou "Viúvo(a)".
    4. Data de nascimento ('data_nascimento') no formato "YYYY-MM-DD". Se disser apenas a idade, estime o ano baseado no ano atual ({datetime.datetime.now().year}), deixando MM-DD como "01-01".
    5. 'first_name' é apenas o primeiro nome. 'last_name' é o restante.

    TEXTO:
    "{texto_livre}"

    Retorne APENAS o JSON no formato:
    {{
        "first_name": "", "last_name": "", "apelido": "", "email": "", "telefone": "",
        "data_nascimento": "", "cpf": "", "rg": "", "sexo": "", "estado_civil": "",
        "profissao": "", "escolaridade": ""
    }}
    """
    try:
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt
        )
        text = response.text.strip()
        if text.startswith("```json"): text = text[7:]
        if text.endswith("```"): text = text[:-3]
        return json.loads(text.strip())
    except Exception as e:
        print(f"Erro no AI Autofill de Membros: {{e}}")
        return {{}}
